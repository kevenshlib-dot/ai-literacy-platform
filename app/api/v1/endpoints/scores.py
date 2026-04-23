"""Scoring and report API endpoints."""
import asyncio
import logging
from io import BytesIO
from uuid import UUID
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select, func, case, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import async_session, get_db
from app.api.deps import get_current_active_user, require_role
from app.models.user import User, Role
from app.models.score import Score, ScoreComplaint
from app.models.answer import AnswerSheet, AnswerSheetStatus
from app.models.exam import Exam
from app.services import score_service
from app.services import diagnostic_service
from app.services import evaluation_service
from app.agents.scoring_agent import multi_model_score
from app.services.report_storage import DIAGNOSTIC_REPORT_KEY, get_report_namespace
from app.services.score_service import serialize_score_detail

router = APIRouter(prefix="/scores", tags=["评分管理"])
logger = logging.getLogger(__name__)

_processing_states: dict[str, dict] = {}
_processing_tasks: dict[str, asyncio.Task] = {}


def _processing_key(answer_sheet_id: UUID | str) -> str:
    return str(answer_sheet_id)


def _processing_payload(
    answer_sheet_id: UUID | str,
    *,
    stage: str,
    score_id: str | None = None,
    diagnostic_ready: bool = False,
    message: str = "",
) -> dict:
    return {
        "answer_sheet_id": str(answer_sheet_id),
        "stage": stage,
        "score_id": score_id,
        "diagnostic_ready": diagnostic_ready,
        "message": message,
    }


def _set_processing_state(
    answer_sheet_id: UUID | str,
    *,
    stage: str,
    score_id: str | None = None,
    diagnostic_ready: bool = False,
    message: str = "",
) -> dict:
    payload = _processing_payload(
        answer_sheet_id,
        stage=stage,
        score_id=score_id,
        diagnostic_ready=diagnostic_ready,
        message=message,
    )
    _processing_states[_processing_key(answer_sheet_id)] = payload
    return payload


def _get_processing_state(answer_sheet_id: UUID | str) -> dict | None:
    return _processing_states.get(_processing_key(answer_sheet_id))


async def _derive_processing_state_from_db(db: AsyncSession, answer_sheet_id: UUID) -> dict:
    sheet = (
        await db.execute(select(AnswerSheet).where(AnswerSheet.id == answer_sheet_id))
    ).scalar_one_or_none()
    if not sheet:
        raise HTTPException(status_code=404, detail="答题卡不存在")

    score = await score_service.get_score_by_sheet(db, answer_sheet_id)
    if not score:
        stage = (
            "submitted"
            if sheet.status == AnswerSheetStatus.SUBMITTED
            else "completed" if sheet.status == AnswerSheetStatus.SCORED else "submitted"
        )
        message = "答卷已提交，等待开始处理。" if stage == "submitted" else "评分已完成。"
        return _processing_payload(answer_sheet_id, stage=stage, message=message)

    diagnostic_ready = bool(get_report_namespace(score, DIAGNOSTIC_REPORT_KEY))
    if diagnostic_ready:
        return _processing_payload(
            answer_sheet_id,
            stage="completed",
            score_id=str(score.id),
            diagnostic_ready=True,
            message="诊断报告已生成。",
        )

    return _processing_payload(
        answer_sheet_id,
        stage="generating_diagnostic",
        score_id=str(score.id),
        diagnostic_ready=False,
        message="成绩已生成，正在生成诊断报告。",
    )


async def _run_answer_sheet_processing(answer_sheet_id: UUID) -> None:
    key = _processing_key(answer_sheet_id)
    try:
        async with async_session() as db:
            score = await score_service.get_score_by_sheet(db, answer_sheet_id, load_details=True)
            if score and get_report_namespace(score, DIAGNOSTIC_REPORT_KEY):
                _set_processing_state(
                    answer_sheet_id,
                    stage="completed",
                    score_id=str(score.id),
                    diagnostic_ready=True,
                    message="诊断报告已生成。",
                )
                await db.commit()
                return

            if not score:
                _set_processing_state(
                    answer_sheet_id,
                    stage="scoring",
                    message="正在评分，请稍候。",
                )
                score = await score_service.score_answer_sheet(db, answer_sheet_id)

            _set_processing_state(
                answer_sheet_id,
                stage="generating_diagnostic",
                score_id=str(score.id),
                message="正在生成诊断报告。",
            )
            await score_service.generate_report(db, score.id, force_refresh=True)
            await diagnostic_service.generate_diagnostic_report(db, score.id, force_refresh=True)
            await db.commit()
            _set_processing_state(
                answer_sheet_id,
                stage="completed",
                score_id=str(score.id),
                diagnostic_ready=True,
                message="诊断报告已生成。",
            )
    except Exception as exc:
        logger.exception("Answer sheet processing failed for %s", answer_sheet_id)
        _set_processing_state(
            answer_sheet_id,
            stage="failed",
            message=f"处理失败：{exc}",
        )
    finally:
        _processing_tasks.pop(key, None)


class PanelScoreRequest(BaseModel):
    """Request for multi-model panel scoring of a subjective answer."""
    stem: str
    correct_answer: str
    student_answer: str
    question_type: str = "short_answer"
    max_score: float = 10.0
    rubric: Optional[dict] = None
    num_evaluators: int = Field(default=3, ge=2, le=5)


class TrainingGenerateRequest(BaseModel):
    """Request to generate training questions based on weak areas."""
    wrong_questions: list[dict] = Field(description="Wrong question data with stem, dimension, etc.")
    count: int = Field(default=5, ge=1, le=10)
    difficulty: int = Field(default=3, ge=1, le=5)


def _ensure_sheet_access(sheet: AnswerSheet | None, current_user: User) -> None:
    if not sheet:
        raise HTTPException(status_code=404, detail="答题卡不存在")
    if sheet.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能处理自己的答题卡")


@router.post("/process/{answer_sheet_id}")
async def start_score_processing(
    answer_sheet_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    sheet = (
        await db.execute(select(AnswerSheet).where(AnswerSheet.id == answer_sheet_id))
    ).scalar_one_or_none()
    _ensure_sheet_access(sheet, current_user)

    derived = await _derive_processing_state_from_db(db, answer_sheet_id)
    if derived["stage"] == "completed":
        return derived

    existing = _processing_tasks.get(_processing_key(answer_sheet_id))
    if existing and not existing.done():
        return _get_processing_state(answer_sheet_id) or derived

    state = _set_processing_state(
        answer_sheet_id,
        stage="submitted",
        score_id=derived.get("score_id"),
        message="答卷已提交，处理中。",
    )
    _processing_tasks[_processing_key(answer_sheet_id)] = asyncio.create_task(
        _run_answer_sheet_processing(answer_sheet_id)
    )
    return state


@router.get("/process/{answer_sheet_id}")
async def get_score_processing_status(
    answer_sheet_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    sheet = (
        await db.execute(select(AnswerSheet).where(AnswerSheet.id == answer_sheet_id))
    ).scalar_one_or_none()
    _ensure_sheet_access(sheet, current_user)

    state = _get_processing_state(answer_sheet_id)
    if state:
        return state
    return await _derive_processing_state_from_db(db, answer_sheet_id)


@router.get("/leaderboard")
async def get_leaderboard(
    limit: int = Query(20, ge=1, le=50),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get leaderboard: top users by best score ratio across all exams."""
    # Subquery: for each user, find their best score ratio
    best_score_subq = (
        select(
            AnswerSheet.user_id,
            func.max(Score.total_score / Score.max_score).label("best_ratio"),
        )
        .join(Score, Score.answer_sheet_id == AnswerSheet.id)
        .where(AnswerSheet.status == AnswerSheetStatus.SCORED)
        .where(AnswerSheet.is_deleted == False)
        .group_by(AnswerSheet.user_id)
        .subquery()
    )

    # Main query: join with users + get the actual score details for the best ratio
    # First get qualifying user_ids with their best ratio
    user_best = (
        select(
            best_score_subq.c.user_id,
            best_score_subq.c.best_ratio,
        )
        .join(User, User.id == best_score_subq.c.user_id)
        .where(
            User.is_deleted == False,
            User.show_on_leaderboard == True,
        )
        .order_by(best_score_subq.c.best_ratio.desc())
        .limit(limit)
        .subquery()
    )

    # Get the full details: join back to get score, exam info
    stmt = (
        select(
            User.id,
            User.username,
            User.full_name,
            Score.total_score,
            Score.max_score,
            (Score.total_score / Score.max_score).label("score_ratio"),
            Score.level,
            Exam.title.label("exam_title"),
            Score.scored_at,
        )
        .join(user_best, User.id == user_best.c.user_id)
        .join(AnswerSheet, and_(
            AnswerSheet.user_id == User.id,
            AnswerSheet.status == AnswerSheetStatus.SCORED,
            AnswerSheet.is_deleted == False,
        ))
        .join(Score, Score.answer_sheet_id == AnswerSheet.id)
        .join(Exam, Exam.id == AnswerSheet.exam_id)
        .where(
            (Score.total_score / Score.max_score) == user_best.c.best_ratio,
        )
        .order_by(
            user_best.c.best_ratio.desc(),
            Score.scored_at.asc(),
        )
    )

    result = await db.execute(stmt)
    rows = result.all()

    # Deduplicate: keep only first row per user (best ratio, earliest time)
    seen_users = set()
    items = []
    rank = 0
    for row in rows:
        uid = str(row.id)
        if uid in seen_users:
            continue
        seen_users.add(uid)
        rank += 1
        items.append({
            "rank": rank,
            "user_id": uid,
            "username": row.username,
            "full_name": row.full_name,
            "total_score": row.total_score,
            "max_score": row.max_score,
            "score_ratio": round(row.score_ratio * 100, 1),
            "level": row.level,
            "exam_title": row.exam_title,
            "scored_at": row.scored_at.isoformat() if row.scored_at else None,
        })
        if len(items) >= limit:
            break

    return {"items": items, "total": len(items)}


@router.post("/leaderboard/opt-out")
async def toggle_leaderboard_visibility(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Toggle current user's leaderboard visibility."""
    current_user.show_on_leaderboard = not current_user.show_on_leaderboard
    await db.commit()
    return {
        "show_on_leaderboard": current_user.show_on_leaderboard,
        "message": "已更新排行榜显示设置",
    }


class ExportScoresRequest(BaseModel):
    user_ids: list[str] = Field(default_factory=list, description="Empty list = export all")


@router.get("/all")
async def list_all_scores(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    keyword: Optional[str] = Query(None),
    username: Optional[str] = Query(None),
    full_name: Optional[str] = Query(None),
    exam_title: Optional[str] = Query(None),
    level: Optional[str] = Query(None),
    sort_field: Optional[str] = Query(None),
    sort_order: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "organizer"])),
):
    """List all users' scored results (admin/organizer only)."""
    query = (
        select(
            User.id.label("user_id"),
            User.username,
            User.full_name,
            Exam.title.label("exam_title"),
            Score.total_score,
            Score.max_score,
            Score.level,
            Score.id.label("score_id"),
            AnswerSheet.submit_time,
            AnswerSheet.id.label("answer_sheet_id"),
        )
        .select_from(Score)
        .join(AnswerSheet, Score.answer_sheet_id == AnswerSheet.id)
        .join(User, AnswerSheet.user_id == User.id)
        .join(Exam, AnswerSheet.exam_id == Exam.id)
        .where(User.is_deleted == False)
        .where(AnswerSheet.is_deleted == False)
    )
    count_q = (
        select(func.count(Score.id))
        .select_from(Score)
        .join(AnswerSheet, Score.answer_sheet_id == AnswerSheet.id)
        .join(User, AnswerSheet.user_id == User.id)
        .join(Exam, AnswerSheet.exam_id == Exam.id)
        .where(User.is_deleted == False)
        .where(AnswerSheet.is_deleted == False)
    )

    if keyword:
        kw = f"%{keyword}%"
        kw_filter = User.username.ilike(kw) | User.full_name.ilike(kw) | Exam.title.ilike(kw)
        query = query.where(kw_filter)
        count_q = count_q.where(kw_filter)

    # Per-field filters
    if username:
        uname_f = User.username.ilike(f"%{username}%")
        query = query.where(uname_f)
        count_q = count_q.where(uname_f)
    if full_name:
        fname_f = User.full_name.ilike(f"%{full_name}%")
        query = query.where(fname_f)
        count_q = count_q.where(fname_f)
    if exam_title:
        etitle_f = Exam.title.ilike(f"%{exam_title}%")
        query = query.where(etitle_f)
        count_q = count_q.where(etitle_f)
    if level:
        # Support comma-separated levels for multi-select filter
        levels = [lv.strip() for lv in level.split(",") if lv.strip()]
        if levels:
            query = query.where(Score.level.in_(levels))
            count_q = count_q.where(Score.level.in_(levels))

    # Sorting
    sort_map = {
        "total_score": Score.total_score,
        "score_ratio": (Score.total_score / Score.max_score),
        "submit_time": AnswerSheet.submit_time,
        "username": User.username,
        "full_name": User.full_name,
        "exam_title": Exam.title,
    }
    order_col = sort_map.get(sort_field, AnswerSheet.submit_time)
    if sort_order == "ascend":
        query = query.order_by(order_col.asc())
    else:
        query = query.order_by(order_col.desc())

    total = (await db.execute(count_q)).scalar() or 0
    result = await db.execute(query.offset(skip).limit(limit))
    rows = result.all()

    items = []
    for row in rows:
        ratio = round(row.total_score / row.max_score * 100, 1) if row.max_score else 0
        items.append({
            "user_id": str(row.user_id),
            "username": row.username,
            "full_name": row.full_name,
            "exam_title": row.exam_title,
            "total_score": row.total_score,
            "max_score": row.max_score,
            "score_ratio": ratio,
            "level": row.level,
            "score_id": str(row.score_id),
            "answer_sheet_id": str(row.answer_sheet_id),
            "submit_time": row.submit_time.isoformat() if row.submit_time else None,
        })

    return {"total": total, "items": items}


@router.post("/export")
async def export_scores(
    body: ExportScoresRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "organizer"])),
):
    """Export scores to XLSX file (admin/organizer only)."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    query = (
        select(
            User.username,
            User.full_name,
            Exam.title.label("exam_title"),
            Score.total_score,
            Score.max_score,
            Score.level,
            AnswerSheet.submit_time,
        )
        .select_from(Score)
        .join(AnswerSheet, Score.answer_sheet_id == AnswerSheet.id)
        .join(User, AnswerSheet.user_id == User.id)
        .join(Exam, AnswerSheet.exam_id == Exam.id)
        .where(User.is_deleted == False)
        .where(AnswerSheet.is_deleted == False)
        .order_by(User.username, AnswerSheet.submit_time.desc())
    )

    if body.user_ids:
        import uuid as _uuid
        uid_list = [_uuid.UUID(uid) for uid in body.user_ids]
        query = query.where(User.id.in_(uid_list))

    result = await db.execute(query)
    rows = result.all()

    # Build XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成绩导出"

    headers = ["用户名", "姓名", "考试名称", "得分", "满分", "得分率", "等级", "提交时间"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        ratio = round(row.total_score / row.max_score * 100, 1) if row.max_score else 0
        submit_str = row.submit_time.strftime("%Y-%m-%d %H:%M:%S") if row.submit_time else ""
        values = [
            row.username,
            row.full_name or "",
            row.exam_title,
            row.total_score,
            row.max_score,
            f"{ratio}%",
            row.level or "",
            submit_str,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in (4, 5):
                cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                length = len(str(cell.value or ""))
                if length > max_len:
                    max_len = length
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=scores_export.xlsx"},
    )


@router.delete("/{answer_sheet_id}")
async def delete_score(
    answer_sheet_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Soft-delete own score (any user can delete their own scores only)."""
    sheet = (await db.execute(
        select(AnswerSheet).where(AnswerSheet.id == answer_sheet_id)
    )).scalar_one_or_none()
    if not sheet:
        raise HTTPException(status_code=404, detail="成绩不存在")
    if sheet.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能删除自己的成绩")
    if sheet.is_deleted:
        raise HTTPException(status_code=400, detail="该成绩已被删除")

    sheet.is_deleted = True
    sheet.deleted_at = datetime.now(timezone.utc)
    await db.commit()
    return {"message": "成绩已删除"}


@router.get("/archived")
async def list_archived_scores(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin"])),
):
    """List soft-deleted scores (admin only, for archive/statistics)."""
    query = (
        select(
            User.id.label("user_id"),
            User.username,
            User.full_name,
            Exam.title.label("exam_title"),
            Score.total_score,
            Score.max_score,
            Score.level,
            Score.id.label("score_id"),
            AnswerSheet.submit_time,
            AnswerSheet.id.label("answer_sheet_id"),
            AnswerSheet.deleted_at,
        )
        .select_from(Score)
        .join(AnswerSheet, Score.answer_sheet_id == AnswerSheet.id)
        .join(User, AnswerSheet.user_id == User.id)
        .join(Exam, AnswerSheet.exam_id == Exam.id)
        .where(AnswerSheet.is_deleted == True)
    )
    count_q = (
        select(func.count(Score.id))
        .select_from(Score)
        .join(AnswerSheet, Score.answer_sheet_id == AnswerSheet.id)
        .where(AnswerSheet.is_deleted == True)
    )

    total = (await db.execute(count_q)).scalar() or 0
    result = await db.execute(
        query.order_by(AnswerSheet.deleted_at.desc()).offset(skip).limit(limit)
    )
    rows = result.all()

    items = []
    for row in rows:
        ratio = round(row.total_score / row.max_score * 100, 1) if row.max_score else 0
        items.append({
            "user_id": str(row.user_id),
            "username": row.username,
            "full_name": row.full_name,
            "exam_title": row.exam_title,
            "total_score": row.total_score,
            "max_score": row.max_score,
            "score_ratio": ratio,
            "level": row.level,
            "score_id": str(row.score_id),
            "answer_sheet_id": str(row.answer_sheet_id),
            "submit_time": row.submit_time.isoformat() if row.submit_time else None,
            "deleted_at": row.deleted_at.isoformat() if row.deleted_at else None,
        })

    return {"total": total, "items": items}


@router.post("/{answer_sheet_id}/restore")
async def restore_score(
    answer_sheet_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin"])),
):
    """Restore a soft-deleted score (admin only)."""
    sheet = (await db.execute(
        select(AnswerSheet).where(AnswerSheet.id == answer_sheet_id)
    )).scalar_one_or_none()
    if not sheet:
        raise HTTPException(status_code=404, detail="成绩不存在")
    if not sheet.is_deleted:
        raise HTTPException(status_code=400, detail="该成绩未被删除")

    sheet.is_deleted = False
    sheet.deleted_at = None
    await db.commit()
    return {"message": "成绩已恢复"}


@router.post("/panel-score")
async def panel_score_answer(
    body: PanelScoreRequest,
    current_user: User = Depends(require_role(["admin", "organizer", "reviewer"])),
):
    """Score a subjective answer using multi-model evaluator panel."""
    result = multi_model_score(
        stem=body.stem,
        correct_answer=body.correct_answer,
        student_answer=body.student_answer,
        question_type=body.question_type,
        max_score=body.max_score,
        rubric=body.rubric,
        num_evaluators=body.num_evaluators,
    )
    return result


@router.post("/grade/{sheet_id}")
async def grade_answer_sheet(
    sheet_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "organizer", "reviewer"])),
):
    """Score a submitted answer sheet."""
    try:
        score = await score_service.score_answer_sheet(db, sheet_id)
        await score_service.generate_report(db, score.id, force_refresh=True)
        await diagnostic_service.generate_diagnostic_report(db, score.id, force_refresh=True)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return {
        "score_id": str(score.id),
        "total_score": score.total_score,
        "max_score": score.max_score,
        "level": score.level,
        "dimension_scores": score.dimension_scores,
    }


@router.get("/sheet/{sheet_id}")
async def get_score_by_sheet(
    sheet_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get score for an answer sheet."""
    score = await score_service.get_score_by_sheet(db, sheet_id, load_details=True)
    if not score:
        raise HTTPException(status_code=404, detail="成绩不存在")
    return {
        "score_id": str(score.id),
        "total_score": score.total_score,
        "max_score": score.max_score,
        "level": score.level,
        "dimension_scores": score.dimension_scores,
        "details": [
            serialize_score_detail(d)
            for d in score.details
        ],
    }


@router.post("/training/generate")
async def generate_training_questions(
    body: TrainingGenerateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Generate training questions: try LLM first, fall back to DB if slow/failed."""
    import asyncio
    import logging
    from uuid import UUID as PyUUID
    from sqlalchemy import select, func
    from app.models.question import Question
    from app.agents.question_agent import generate_questions_via_llm

    logger = logging.getLogger(__name__)

    # Collect metadata from wrong questions
    stems = [q.get("stem", "") for q in body.wrong_questions if q.get("stem")]
    dimensions = list(set(q.get("dimension", "") for q in body.wrong_questions if q.get("dimension")))
    wrong_qids = [q.get("question_id") for q in body.wrong_questions if q.get("question_id")]
    tags = []
    for q in body.wrong_questions:
        tags.extend(q.get("knowledge_tags") or [])
    unique_tags = list(set(tags))[:10]

    # Only objective types for client-side scoring
    qtypes = list(set(q.get("question_type", "single_choice") for q in body.wrong_questions))
    objective_types = {"single_choice", "multiple_choice", "true_false"}
    valid_types = [t for t in qtypes if t in objective_types]
    if not valid_types:
        valid_types = ["single_choice"]

    # Strategy 1: Try LLM generation (with timeout)
    questions = []
    try:
        content = "\n".join(stems)
        custom_prompt = (
            f"以下是学生答错的题目，请根据这些题目涉及的知识点，"
            f"生成{body.count}道类似但不完全相同的变体题目，用于针对性训练。"
        )
        if dimensions:
            custom_prompt += f"涉及维度：{'、'.join(dimensions)}。"
        if unique_tags:
            custom_prompt += f"相关知识标签：{'、'.join(unique_tags)}。"
        custom_prompt += "要求：题目应覆盖相同知识点但变换考查角度或情境。"

        questions = await asyncio.wait_for(
            asyncio.to_thread(
                generate_questions_via_llm,
                content=content,
                question_types=valid_types,
                count=body.count,
                difficulty=body.difficulty,
                custom_prompt=custom_prompt,
            ),
            timeout=60,
        )
    except (asyncio.TimeoutError, Exception) as e:
        logger.warning(f"LLM training generation failed/timeout: {e}, falling back to DB")

    # Strategy 2: Fall back to DB if LLM returned nothing
    if not questions:
        logger.info("Falling back to DB question pool for training")
        # Build query: same dimensions, objective types, exclude the wrong questions themselves
        stmt = select(Question).where(
            Question.question_type.in_(valid_types),
        )
        if dimensions:
            stmt = stmt.where(Question.dimension.in_(dimensions))
        if wrong_qids:
            try:
                exclude_ids = [PyUUID(qid) for qid in wrong_qids]
                stmt = stmt.where(Question.id.notin_(exclude_ids))
            except (ValueError, AttributeError):
                pass
        stmt = stmt.order_by(func.random()).limit(body.count)
        result = await db.execute(stmt)
        db_questions = result.scalars().all()

        for q in db_questions:
            qtype = q.question_type.value if hasattr(q.question_type, 'value') else q.question_type
            questions.append({
                "question_type": qtype,
                "stem": q.stem,
                "options": q.options,
                "correct_answer": q.correct_answer,
                "explanation": q.explanation,
                "knowledge_tags": q.knowledge_tags,
            })

    return {"questions": questions, "count": len(questions)}


# ── Complaint endpoints (must be before /{score_id} parameterized routes) ──

class ComplaintRequest(BaseModel):
    score_detail_id: UUID
    reason: str = Field(min_length=1, max_length=2000)


class HandleComplaintRequest(BaseModel):
    status: str  # "accepted" or "rejected"
    reply: str = Field(default="", max_length=2000)


@router.post("/complaints")
async def submit_complaint(
    body: ComplaintRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Submit a complaint/feedback on a specific question's score."""
    try:
        complaint = await score_service.create_complaint(
            db, body.score_detail_id, current_user.id, body.reason
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return {"id": str(complaint.id), "message": "投诉已提交，我们会尽快处理"}


@router.get("/complaints")
async def list_complaints(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    status: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "organizer"])),
):
    """List all complaints (admin/organizer only)."""
    items, total = await score_service.list_complaints(db, skip, limit, status)
    return {"total": total, "items": items}


@router.put("/complaints/{complaint_id}")
async def handle_complaint(
    complaint_id: UUID,
    body: HandleComplaintRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "organizer"])),
):
    """Handle a complaint: accept or reject with reply."""
    try:
        complaint = await score_service.handle_complaint(
            db, complaint_id, body.status, body.reply
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return {"id": str(complaint.id), "status": complaint.status.value, "message": "处理完成"}


@router.get("/{score_id}/review")
async def get_review_data(
    score_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get wrong answer details for review/复盘."""
    try:
        items = await score_service.get_wrong_answer_details(db, score_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return {"score_id": str(score_id), "wrong_items": items, "total_wrong": len(items)}


@router.get("/{score_id}/full-review")
async def get_full_review_data(
    score_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get ALL answer details (correct and wrong) for full grading review."""
    try:
        items = await score_service.get_all_answer_details(db, score_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    correct_count = sum(1 for i in items if i.get("is_correct"))
    wrong_count = sum(1 for i in items if i.get("is_correct") is False)
    return {
        "score_id": str(score_id),
        "items": items,
        "total": len(items),
        "correct_count": correct_count,
        "wrong_count": wrong_count,
    }


@router.get("/{score_id}")
async def get_score(
    score_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get score by ID."""
    score = await score_service.get_score_by_id(db, score_id, load_details=True)
    if not score:
        raise HTTPException(status_code=404, detail="成绩不存在")
    return {
        "score_id": str(score.id),
        "total_score": score.total_score,
        "max_score": score.max_score,
        "level": score.level,
        "dimension_scores": score.dimension_scores,
        "details": [
            serialize_score_detail(d)
            for d in score.details
        ],
    }


@router.post("/{score_id}/report")
async def generate_report(
    score_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Generate a detailed score report with analysis and recommendations."""
    try:
        report = await score_service.generate_report(db, score_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    await db.commit()
    return report


@router.get("/{score_id}/diagnostic")
async def get_diagnostic_report(
    score_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Generate five-dimensional diagnostic report with radar chart data (T022)."""
    try:
        report = await diagnostic_service.generate_diagnostic_report(db, score_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    await db.commit()
    return report


@router.get("/{score_id}/evaluation")
async def get_evaluation_feedback(
    score_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get evaluation level, ranking, and motivational feedback (T023)."""
    try:
        feedback = await evaluation_service.get_evaluation_feedback(db, score_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    await db.commit()
    return feedback
